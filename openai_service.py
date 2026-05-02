import csv
import json
import re
from io import BytesIO
from pathlib import Path

from openai import OpenAI


OPENAI_KEY_RE = re.compile(r"sk-[A-Za-z0-9_-]+")


def extract_openai_key_from_text(text: str) -> str | None:
    match = OPENAI_KEY_RE.search(text)
    if match:
        return match.group(0)
    return None


def load_openai_key_from_file(path: str | None) -> str | None:
    if not path:
        return None

    key_path = Path(path)
    if not key_path.exists():
        raise FileNotFoundError(f"OpenAI key file not found: {key_path}")

    raw = key_path.read_bytes()
    if raw.startswith(b"PK"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read Excel-formatted OpenAI key files.") from exc

        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is None:
                            continue
                        key = extract_openai_key_from_text(str(cell))
                        if key:
                            return key
        finally:
            workbook.close()
        raise ValueError(f"No OpenAI API key was found in Excel file: {key_path}")

    text = raw.decode("utf-8-sig")
    key = extract_openai_key_from_text(text)
    if key:
        return key

    rows = csv.DictReader(text.splitlines())
    for row in rows:
        for column in ("openai_api_key", "api_key", "api key", "key", "secret key"):
            value = row.get(column) or row.get(column.title())
            if value:
                key = extract_openai_key_from_text(value) or value.strip()
                if key:
                    return key

    raise ValueError(f"No OpenAI API key was found in CSV file: {key_path}")


def request_openai_advice(reading: dict, plant_name: str, model: str, api_key: str | None) -> dict:
    client = OpenAI(api_key=api_key) if api_key else OpenAI()
    prompt = {
        "plant": plant_name,
        "sensor_reading": reading,
        "hardware": {
            "temperature_sensor": "LM35DZ on Pico GP26 / ADC0",
            "soil_moisture_sensor": "Soil moisture analog output on ADS1115 A0, Pico I2C GP0 SDA / GP1 SCL",
            "pump_relay": "GP15 relay1, active-low, 0 means ON",
            "fan_relay": "GP14 relay2, active-low, 0 means ON",
        },
        "task": (
            "Compare the live readings with typical cherry tomato growing needs. "
            "Return conservative automation recommendations. If data is ambiguous, "
            "prefer no relay action and explain the uncertainty. Write all human-facing "
            "report fields in Turkish."
        ),
    }
    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": (
                    "You are an indoor agriculture assistant. Return only valid JSON with keys: "
                    "temperature_assessment, soil_moisture_assessment, recommended_actions, "
                    "user_report, safety_note. recommended_actions must include boolean fan_on "
                    "and pump_on plus short reason strings fan_reason and pump_reason. "
                    "Use Turkish for all explanatory text values."
                ),
            },
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
    )
    text = response.output_text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"OpenAI response was not valid JSON:\n{text}") from exc

    actions = data.setdefault("recommended_actions", {})
    actions["fan_on"] = bool(actions.get("fan_on"))
    actions["pump_on"] = bool(actions.get("pump_on"))
    actions.setdefault("fan_reason", "")
    actions.setdefault("pump_reason", "")
    data.setdefault("user_report", "")
    data.setdefault("safety_note", "")
    return data
